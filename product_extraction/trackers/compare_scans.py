import os
import re
import sys
from pathlib import Path

# Enable direct-script execution (python trackers/compare_scans.py)
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ─── Shared Excel utilities (Unit 3) ───────────────────────────────
from common.excel_utils import (
    GREEN_FILL,
    RED_FILL,
    YELLOW_FILL,
    BLUE_FILL,
    LIGHT_BLUE_FILL,
    BOLD_FONT,
    NORMAL_FONT,
    RED_FONT,
    GREEN_FONT,
    read_excel,
    style_header,
    auto_width,
    set_fill,
)
from common.file_registry import get_file
from common.path_registry import (
    INTERMEDIATE_DIR,
    LEGACY_APP_DIR,
    RUNTIME_CACHE_DIR,
    RUNTIME_REPORTS_DIR,
    get_dated_reports_dir,
)
from common.file_utils import find_first_glob_match, find_latest_dated
from common.price_utils import select_effective_price as _select_effective_price
from common.color_utils import split_color_values as _split_color_values
from common.text_utils import (
    extract_numeric_code as _extract_numeric_code,
    normalize_text as _normalize_text,
)


def normalize_text(s):
    """حذف کاراکترهای Unicode نامرئی (مثل ZWNJ، ZWJ، Zero-Width Space و غیره)"""
    return _normalize_text(s)


def parse_colors(val):
    """پارس رنگ‌ها با حذف کاراکترهای نامرئی Unicode برای مقایسه دقیق"""
    if pd.isna(val) or str(val).strip() == "":
        return set()
    return {
        normalize_text(c)
        for c in _split_color_values(str(val))
        if normalize_text(c)
    }

def extract_code(text):
    """استخراج کد عددی محصول (۳ تا ۶ رقم) از یک متن (مثل sku یا نام محصول)."""
    return _extract_numeric_code(text, min_digits=3, max_digits=6)


def effective_price(regular, sale):
    """قیمت موثر: اگه sale_price داشت اون رو برگردان، وگرنه regular_price."""
    return _select_effective_price(regular, sale)


# ─── توابع پیدا کردن فایل‌ها ──────────────────────────────────────


def find_latest_scan(folder):
    latest = find_latest_dated(
        folder,
        "product_details_*.xlsx",
        r"product_details_(\d{8}_\d{6})\.xlsx$",
    )
    return str(latest) if latest else None


def find_links_file(folder):
    """پیدا کردن فایل extracted_products*.xlsx (لینک محصولات) در یک پوشه."""
    match = find_first_glob_match(folder, f"{Path(get_file('extracted_products')).stem}*.xlsx")
    return str(match) if match else None


def find_latest_woo_file(folder):
    latest = find_latest_dated(
        folder,
        "woocommerce_import_*.csv",
        r"woocommerce_import_(\d{8})_\d+\.csv$",
        recursive=True,
    )
    return str(latest) if latest else None


# ─── بارگذاری WooCommerce ──────────────────────────────────────────


def load_woo(woo_path):
    df = pd.read_csv(woo_path, dtype=str, low_memory=False)
    df.columns = [c.strip() for c in df.columns]
    df["sku"] = df["sku"].fillna("").str.strip()
    df["parent_sku"] = df["parent_sku"].fillna("").str.strip()

    products = df[df["post_type"].str.strip() == "product"].copy()
    variations = df[df["post_type"].str.strip() == "product_variation"].copy()

    products = products.rename(
        columns={
            "post_title": "نام_محصول",
            "regular_price": "قیمت_اصلی",
            "sale_price": "قیمت_فروش",
            "attribute:color": "رنگ",
        }
    )
    return products, variations


# ─── بارگذاری فایل لینک محصولات ─────────────────────────────────────


def load_product_links(links_path):
    """
    فایل extracted_products.xlsx رو می‌خونه (ستون‌ها: No, Product Name, Product URL, Price)
    و یک دیکشنری {کد محصول: {"name":..., "url":..., "price":...}} برمی‌گردونه.
    کد محصول از داخل Product Name استخراج می‌شه (مثلاً «کوله پشتی کد 9266» -> 9266).
    """
    df = read_excel(links_path, dtype=str)
    df.columns = [c.strip() for c in df.columns]

    links = {}
    for _, row in df.iterrows():
        name = row.get("Product Name", "")
        url = row.get("Product URL", "")
        price = row.get("Price", "")
        code = extract_code(name)
        if code:
            links[code] = {"name": name, "url": url, "price": price}
    return links


# ─── نوشتن شیت‌ها ─────────────────────────────────────────────────


def write_sheet_new(wb, new_products, new_colors):
    """
    new_products: محصولات کاملاً جدید (سبز)
    new_colors: محصولات موجود با رنگ جدید (آبی) - فقط رنگ‌های جدید در ستون رنگ
    """
    ws = wb.create_sheet("محصولات جدید")

    # ستون‌های مشترک
    base_cols = list(new_products.columns)
    if "نوع" not in base_cols:
        base_cols = ["نوع"] + base_cols
    ws.append(base_cols)
    style_header(ws, 1, len(base_cols))

    # محصولات کاملاً جدید
    for _, row in new_products.iterrows():
        data = ["محصول جدید"] + [row[c] for c in new_products.columns]
        ws.append(data)
        set_fill(ws, ws.max_row, len(base_cols), GREEN_FILL)
        for col in range(1, len(base_cols) + 1):
            ws.cell(ws.max_row, col).font = NORMAL_FONT

    # محصولات با رنگ جدید
    for row in new_colors:
        ws.append(["رنگ جدید"] + row)
        set_fill(ws, ws.max_row, len(base_cols), LIGHT_BLUE_FILL)
        for col in range(1, len(base_cols) + 1):
            ws.cell(ws.max_row, col).font = NORMAL_FONT

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_sheet_removed(wb, removed_products):
    ws = wb.create_sheet("محصولات حذف‌شده")
    cols = list(removed_products.columns)
    ws.append(cols)
    style_header(ws, 1, len(cols))
    for _, row in removed_products.iterrows():
        ws.append([row[c] for c in cols])
        set_fill(ws, ws.max_row, len(cols), RED_FILL)
        for col in range(1, len(cols) + 1):
            ws.cell(ws.max_row, col).font = NORMAL_FONT
    auto_width(ws)
    ws.freeze_panes = "A2"


def write_sheet_price(wb, price_changes):
    ws = wb.create_sheet("تغییر قیمت")
    headers = ["sku", "نام_محصول", "قیمت قبل", "قیمت بعد", "مقدار تغییر", "درصد تغییر"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in price_changes:
        ws.append(row)
        r = ws.max_row
        try:
            delta = int(str(row[4]).replace(",", "").replace("+", ""))
            is_increase = delta > 0
        except:
            is_increase = None

        if is_increase is True:
            row_fill = PatternFill("solid", start_color="FFCCCC")  # قرمز روشن
            delta_font = RED_FONT
        elif is_increase is False:
            row_fill = PatternFill("solid", start_color="C6EFCE")  # سبز روشن
            delta_font = GREEN_FONT
        else:
            row_fill = YELLOW_FILL
            delta_font = NORMAL_FONT

        set_fill(ws, r, len(headers), row_fill)
        for col in range(1, len(headers) + 1):
            ws.cell(r, col).font = NORMAL_FONT
        # ستون مقدار تغییر و درصد با رنگ خاص
        ws.cell(r, 5).font = delta_font
        ws.cell(r, 6).font = delta_font

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_sheet_colors(wb, color_changes):
    ws = wb.create_sheet("تغییر رنگ‌ها")
    headers = [
        "sku",
        "نام_محصول",
        "رنگ‌های قبل",
        "رنگ‌های بعد",
        "رنگ‌های اضافه‌شده",
        "رنگ‌های حذف‌شده",
        "SKU متغیر حذف‌شده",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in color_changes:
        ws.append(row)
        r = ws.max_row
        added = str(row[4]) if row[4] else ""
        removed = str(row[5]) if row[5] else ""
        if added and removed:
            fill = YELLOW_FILL
        elif added:
            fill = GREEN_FILL
        else:
            fill = RED_FILL
        set_fill(ws, r, len(headers), fill)
        for col in range(1, len(headers) + 1):
            ws.cell(r, col).font = NORMAL_FONT
        if len(row) >= 7 and row[6]:
            ws.cell(r, 7).font = BOLD_FONT
    auto_width(ws)
    ws.freeze_panes = "A2"


def write_summary(wb, counts, woo_name, scan_name):
    ws = wb.active
    ws.title = "خلاصه تغییرات"
    ws.sheet_view.rightToLeft = True

    ws.append(["خلاصه مقایسه تغییرات"])
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 25

    ws.append([])
    ws.append(["فایل مبدا (سایت فعلی):", woo_name, ""])
    ws.append(["فایل مقصد (اسکن جدید):", scan_name, ""])
    for r in [3, 4]:
        ws.cell(r, 1).font = BOLD_FONT
        ws.cell(r, 2).font = NORMAL_FONT

    ws.append([])
    headers = ["نوع تغییر", "تعداد", "توضیح"]
    ws.append(headers)
    style_header(ws, 6, 3)

    rows = [
        (
            "محصولات جدید",
            counts["new"],
            "محصولات کاملاً جدید + محصولات با رنگ جدید",
            GREEN_FILL,
        ),
        (
            "محصولات حذف‌شده",
            counts["removed"],
            "محصولاتی که دیگر در اسکن جدید نیستند",
            RED_FILL,
        ),
        ("تغییر قیمت", counts["price"], "محصولاتی که قیمت‌شان تغییر کرده", YELLOW_FILL),
        (
            "تغییر رنگ‌ها",
            counts["colors"],
            "محصولاتی که رنگ‌شان اضافه/حذف شده",
            BLUE_FILL,
        ),
    ]
    for label, count, desc, fill in rows:
        ws.append([label, count, desc])
        r = ws.max_row
        for col in range(1, 4):
            ws.cell(r, col).fill = fill
            ws.cell(r, col).font = NORMAL_FONT
            ws.cell(r, col).alignment = Alignment(
                horizontal="right" if col != 2 else "center"
            )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 45


def write_sheet_woo_ready(wb, new_products, new_colors_full):
    """
    شیت آماده برای تولید فایل ووکامرس:
    - محصولات کاملاً جدید (سبز) — تمام ستون‌های اسکن
    - محصولات با رنگ جدید (آبی) — تمام ستون‌های اسکن، ستون رنگ = همه رنگ‌های نهایی
    """
    ws = wb.create_sheet("آماده برای ووکامرس")

    # ستون‌ها از df اسکن
    scan_cols = list(new_products.columns)
    ws.append(scan_cols)
    style_header(ws, 1, len(scan_cols))

    # بخش اول: محصولات کاملاً جدید (سبز)
    for _, row in new_products.iterrows():
        ws.append([row[c] for c in scan_cols])
        set_fill(ws, ws.max_row, len(scan_cols), GREEN_FILL)
        for col in range(1, len(scan_cols) + 1):
            ws.cell(ws.max_row, col).font = NORMAL_FONT

    # بخش دوم: محصولات با رنگ جدید (آبی) — رنگ ستون = colors_after کامل
    for row in new_colors_full:
        ws.append([row[c] if c in row.index else "" for c in scan_cols])
        set_fill(ws, ws.max_row, len(scan_cols), LIGHT_BLUE_FILL)
        for col in range(1, len(scan_cols) + 1):
            ws.cell(ws.max_row, col).font = NORMAL_FONT

    auto_width(ws)
    ws.freeze_panes = "A2"


def write_sheet_links(wb, new_products, new_colors_full, links_map):
    """
    شیت «لینک محصولات جدید»:
    برای هر محصول در شیت «آماده برای ووکامرس» (محصولات کاملاً جدید + محصولات با رنگ جدید)،
    بر اساس کد استخراج‌شده از sku، لینک متناظر از links_map رو پیدا می‌کنه.

    new_products: DataFrame محصولات کاملاً جدید (همون چیزی که به write_sheet_woo_ready می‌ره) — رنگ سبز
    new_colors_full: لیستی از Series محصولات موجود با رنگ جدید — رنگ آبی
    links_map: خروجی load_product_links -> {کد: {"name":..., "url":..., "price":...}}

    خروجی: (matched_count, total_count)
    """
    # ساختار این شیت دقیقاً مشابه فایل extracted_products.xlsx است
    # تا برنامه‌ای که بعداً این فایل رو می‌خونه بتونه ستون‌ها رو تشخیص بده.
    ws = wb.create_sheet("لینک محصولات جدید")
    headers = ["No", "Product Name", "Product URL", "Price"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    # همون منطق رنگ شیت «آماده برای ووکامرس»:
    # سبز = محصولات کاملاً جدید | آبی = محصولات موجود با رنگ جدید
    records = [(row, GREEN_FILL) for _, row in new_products.iterrows()] + [
        (row, LIGHT_BLUE_FILL) for row in new_colors_full
    ]

    matched = 0
    no = 0
    for rec, fill in records:
        sku = str(rec.get("sku", "") or "").strip()
        name = rec.get("نام_محصول", "")
        code = extract_code(sku)
        info = links_map.get(code) if code else None

        no += 1
        if info:
            matched += 1
            product_name = info.get("name") or name
            url = info.get("url", "")
            price = info.get("price", "")
        else:
            product_name = name
            url = ""
            price = ""

        ws.append([no, product_name, url, price])

        set_fill(ws, ws.max_row, len(headers), fill)
        for col in range(1, len(headers) + 1):
            ws.cell(ws.max_row, col).font = NORMAL_FONT

    auto_width(ws)
    ws.freeze_panes = "A2"
    return matched, len(records)


# ─── مقایسه اصلی ──────────────────────────────────────────────────


def compare(scan_path, woo_path, output_path, links_path=None):
    """Compare a new scan against the previous WooCommerce catalog.

    Writes the Excel change report and three CSV manifests (new /
    updated / image-subset) into the output folder, then returns a
    structured partition of SKUs for programmatic callers:

        {
          'new_skus':          [...],   # in scan, not in catalog
          'changed_skus':      [...],   # price and/or color changed
          'price_changed_skus':[...],
          'color_added_skus':  [...],   # gained a new color variant
          'image_subset_skus': [...],   # new ∪ color_added (images to (re)process)
          'removed_skus':      [...],
          'counts':            {...},
          'manifests':         {'new':..., 'updated':..., 'image_subset':...},
        }
    """
    df_scan = read_excel(scan_path, dtype={"sku": str})
    df_scan["sku"] = df_scan["sku"].str.strip()

    woo_products, woo_variations = load_woo(woo_path)
    woo_products["sku"] = woo_products["sku"].str.strip()

    # lookup: (parent_sku, رنگ نرمال‌شده) -> variation_sku
    var_lookup = {}
    for _, row in woo_variations.iterrows():
        parent = normalize_text(str(row.get("parent_sku", "") or ""))
        color = normalize_text(str(row.get("attribute:color", "") or ""))
        vsku = str(row.get("sku", "") or "").strip()
        if parent and color and vsku:
            var_lookup[(parent, color)] = vsku

    skus_woo = set(woo_products["sku"])
    skus_scan = set(df_scan["sku"])

    # ── محصولات جدید ──────────────────────────────────────────────
    new_skus = skus_scan - skus_woo
    new_products = df_scan[df_scan["sku"].isin(new_skus)].copy()

    # ── محصولات حذف‌شده ────────────────────────────────────────────
    removed_skus = skus_woo - skus_scan
    removed_products = woo_products[woo_products["sku"].isin(removed_skus)].copy()
    removed_products = removed_products[
        ["sku", "نام_محصول", "قیمت_اصلی", "قیمت_فروش", "رنگ"]
    ].copy()

    def get_var_skus(parent):
        skus = [v for (p, c), v in var_lookup.items() if p == parent]
        return " | ".join(skus) if skus else ""

    removed_products["SKU متغیرها"] = removed_products["sku"].apply(get_var_skus)

    # ── محصولات مشترک ─────────────────────────────────────────────
    common_skus = skus_woo & skus_scan
    d_woo = (
        woo_products[woo_products["sku"].isin(common_skus)]
        .drop_duplicates("sku")
        .set_index("sku")
    )
    d_scan = (
        df_scan[df_scan["sku"].isin(common_skus)]
        .drop_duplicates("sku")
        .set_index("sku")
    )

    # ── تغییر قیمت ────────────────────────────────────────────────
    price_changes = []
    for sku in common_skus:
        r_woo = d_woo.loc[sku]
        r_scan = d_scan.loc[sku]

        p_before = effective_price(r_woo.get("قیمت_اصلی"), r_woo.get("قیمت_فروش"))
        p_after = effective_price(r_scan.get("قیمت_اصلی"), r_scan.get("قیمت_فروش"))

        if p_before is None or p_after is None:
            continue
        if p_before == p_after:
            continue

        diff = p_after - p_before
        diff_str = f"+{diff:,}" if diff > 0 else f"{diff:,}"
        pct = (diff / p_before) * 100
        pct_str = f"+{pct:.1f}%" if pct > 0 else f"{pct:.1f}%"

        name = r_scan.get("نام_محصول", "")
        price_changes.append(
            [sku, name, f"{p_before:,}", f"{p_after:,}", diff_str, pct_str]
        )

    # ── تغییر رنگ‌ها ──────────────────────────────────────────────
    color_changes = []
    new_colors = []  # محصولات موجود با رنگ جدید برای شیت محصولات جدید (خلاصه)
    new_colors_full = []  # همان محصولات با تمام ستون‌های اسکن برای شیت ووکامرس
    for sku in common_skus:
        r_woo = d_woo.loc[sku]
        r_scan = d_scan.loc[sku]
        # parse_colors اکنون کاراکترهای نامرئی را حذف می‌کند
        colors_before = parse_colors(r_woo.get("رنگ"))
        colors_after = parse_colors(r_scan.get("رنگ"))
        if colors_before == colors_after:
            continue
        added = " | ".join(sorted(colors_after - colors_before)) or ""
        removed = " | ".join(sorted(colors_before - colors_after)) or ""
        removed_var_skus = [
            var_lookup[(sku, c)]
            for c in sorted(colors_before - colors_after)
            if (sku, c) in var_lookup
        ]
        name = r_scan.get("نام_محصول", "")
        color_changes.append(
            [
                sku,
                name,
                " | ".join(sorted(colors_before)),
                " | ".join(sorted(colors_after)),
                added,
                removed,
                " | ".join(removed_var_skus),
            ]
        )
        # اگه رنگ جدید اضافه شده، به لیست محصولات جدید هم اضافه کن
        if added:
            new_colors.append([sku, name, added])
            # ردیف کامل اسکن با رنگ نهایی (colors_after) برای شیت آماده ووکامرس
            full_row = df_scan[df_scan["sku"] == sku].iloc[0].copy()
            full_row["رنگ"] = " | ".join(sorted(colors_after))
            new_colors_full.append(full_row)

    # ── فایل متنی SKUهای حذف‌شده ──────────────────────────────────
    skus_to_delete = []
    # محصولات کاملاً حذف‌شده: SKU اصلی + همه متغیرهاش
    for _, row in removed_products.iterrows():
        skus_to_delete.append(str(row["sku"]))
        for vsku in str(row.get("SKU متغیرها", "") or "").split("|"):
            v = vsku.strip()
            if v:
                skus_to_delete.append(v)
    # متغیرهای رنگ حذف‌شده از محصولات موجود
    for row in color_changes:
        for vsku in str(row[6] or "").split("|"):
            v = vsku.strip()
            if v:
                skus_to_delete.append(v)

    txt_path = output_path.replace(".xlsx", "_skus_to_delete.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        for s in skus_to_delete:
            f.write(s + "\n")

    # ── بارگذاری فایل لینک محصولات (در صورت وجود) ──────────────────
    links_map = {}
    if links_path and os.path.exists(links_path):
        links_map = load_product_links(links_path)

    # ── ساخت فایل اکسل ────────────────────────────────────────────
    wb = Workbook()
    write_summary(
        wb,
        {
            "new": len(new_products) + len(new_colors),
            "removed": len(removed_products),
            "price": len(price_changes),
            "colors": len(color_changes),
        },
        os.path.basename(woo_path),
        os.path.basename(scan_path),
    )

    write_sheet_new(wb, new_products, new_colors)
    write_sheet_woo_ready(wb, new_products, new_colors_full)

    links_matched = links_total = 0
    if links_path and os.path.exists(links_path):
        links_matched, links_total = write_sheet_links(
            wb, new_products, new_colors_full, links_map
        )

    write_sheet_removed(wb, removed_products)
    write_sheet_price(wb, price_changes)
    write_sheet_colors(wb, color_changes)

    wb.save(output_path)

    # ── Build the SKU → product-page-URL map so downstream image download can
    # fetch only the products it needs. links_map is keyed by numeric code
    # (extract_code), so we resolve each SKU's code to its URL.
    def _url_for_sku(sku):
        code = extract_code(str(sku))
        info = links_map.get(code) if code else None
        return (info.get("url", "") if info else "")

    # ── SKU partition sets (the contract change-detection consumes) ──────
    new_sku_set = {str(s).strip() for s in new_products["sku"]} if "sku" in new_products.columns else set()
    price_changed_skus = {str(r[0]).strip() for r in price_changes}
    # Color-changed = existing products that gained at least one NEW color
    # (added non-empty). These are the only "changed" products whose images
    # must be (re)downloaded; a pure price change reuses existing images.
    color_added_skus = {str(r[0]).strip() for r in color_changes if str(r[4] or "").strip()}
    changed_sku_set = price_changed_skus | color_added_skus
    # Images to (re)process = brand-new products ∪ products with a new color.
    image_subset_skus = new_sku_set | color_added_skus

    # ── Write manifests (new / updated / image-subset) into reports folder ─
    new_manifest = Path(output_path).parent / "new_products_list.csv"
    updated_manifest = Path(output_path).parent / "updated_products_list.csv"
    image_manifest = Path(output_path).parent / "image_subset_list.csv"
    try:
        out_dir = Path(output_path).parent

        # scan sku -> image_urls (if the scan carries them)
        scan_img_map = {}
        if not df_scan.empty and 'sku' in df_scan.columns and 'image_urls' in df_scan.columns:
            for _, r in df_scan.iterrows():
                key = str(r.get('sku', '')).strip()
                if key:
                    scan_img_map[key] = r.get('image_urls', '')

        # New products manifest — always carries sku + resolved product URL so
        # the downloader can fetch by page. Empty (header-only) when none.
        df_new_man = pd.DataFrame()
        sku_col = 'sku' if 'sku' in new_products.columns else ('SKU' if 'SKU' in new_products.columns else None)
        if sku_col:
            df_new_man['sku'] = new_products[sku_col].astype(str).str.strip()
        else:
            df_new_man['sku'] = pd.Series(dtype=str)
        if 'نام_محصول' in new_products.columns:
            df_new_man['name'] = new_products['نام_محصول'].values
        if 'قیمت_اصلی' in new_products.columns:
            df_new_man['price'] = new_products['قیمت_اصلی'].values
        df_new_man['url'] = df_new_man['sku'].map(_url_for_sku).fillna('')
        if scan_img_map:
            df_new_man['image_urls'] = df_new_man['sku'].map(scan_img_map).fillna('')
        df_new_man.to_csv(new_manifest, index=False, encoding='utf-8-sig')

        # Updated products manifest (price + color changes) with change_type.
        updated_rows = []
        for row in price_changes:
            try:
                updated_rows.append({
                    'sku': str(row[0]).strip(), 'name': row[1],
                    'change_type': 'price', 'details': f"{row[2]} -> {row[3]}",
                    'url': _url_for_sku(row[0]),
                })
            except Exception:
                continue
        for row in color_changes:
            try:
                updated_rows.append({
                    'sku': str(row[0]).strip(), 'name': row[1],
                    'change_type': 'color',
                    'details': f"added: {row[4]}; removed: {row[5]}",
                    'url': _url_for_sku(row[0]),
                })
            except Exception:
                continue
        # Always write the file (header-only when empty) so downstream code can
        # rely on its existence rather than guessing.
        pd.DataFrame(updated_rows, columns=['sku', 'name', 'change_type', 'details', 'url']).to_csv(
            updated_manifest, index=False, encoding='utf-8-sig'
        )

        # Image-subset manifest — the exact products whose images must be
        # downloaded/processed/renamed this run (new ∪ color-added). Carries a
        # resolved product URL per SKU so the downloader fetches only these.
        img_rows = []
        for sku in sorted(image_subset_skus):
            img_rows.append({
                'sku': sku,
                'url': _url_for_sku(sku),
                'image_urls': scan_img_map.get(sku, ''),
                'reason': 'new' if sku in new_sku_set else 'color',
            })
        pd.DataFrame(img_rows, columns=['sku', 'url', 'image_urls', 'reason']).to_csv(
            image_manifest, index=False, encoding='utf-8-sig'
        )
    except Exception as e:
        print(f"⚠️  Could not write one or more manifests: {e}")

    print(f"✓ Output saved:      {output_path}")
    print(f"✓ SKUs to delete:    {txt_path}")
    print(f"  New products:      {len(new_products)}")
    print(f"  Removed products:  {len(removed_products)}")
    print(f"  Price changes:     {len(price_changes)}")
    print(f"  Color changes:     {len(color_changes)}")
    print(f"  SKUs to delete:    {len(skus_to_delete)}")
    if links_path and os.path.exists(links_path):
        print(f"  Links matched:     {links_matched} / {links_total}")
    else:
        print(f"  Links file:        not found, sheet skipped")

    # Structured partition returned to programmatic callers (change detection in
    # the auto pipeline). File outputs above stay identical for standalone use.
    return {
        'new_skus': sorted(new_sku_set),
        'changed_skus': sorted(changed_sku_set),
        'price_changed_skus': sorted(price_changed_skus),
        'color_added_skus': sorted(color_added_skus),
        'image_subset_skus': sorted(image_subset_skus),
        'removed_skus': sorted({str(s).strip() for s in removed_products['sku']}) if 'sku' in removed_products.columns else [],
        'counts': {
            'new': len(new_sku_set),
            'changed': len(changed_sku_set),
            'price': len(price_changed_skus),
            'color': len(color_added_skus),
            'removed': len(removed_products),
        },
        'manifests': {
            'new': str(new_manifest),
            'updated': str(updated_manifest),
            'image_subset': str(image_manifest),
        },
    }


# ─── نقطه ورود قابل فراخوانی (برای main.py) ────────────────────────


def find_previous_woo_file():
    """Locate the most recent previously-generated WooCommerce catalog CSV.

    Returns the path string, or None when no prior catalog exists (e.g. the
    very first run). main.py uses a None result to mean "no baseline to compare
    against → run full".
    """
    woo_dirs = [str(RUNTIME_CACHE_DIR / "import_builder" / "uploads")]
    for d in woo_dirs:
        woo_path = find_latest_woo_file(d)
        if woo_path:
            return woo_path
    return None


def run_auto_compare(scan_path=None, woo_path=None, links_path=None):
    """Run change-detection using auto-located files and return the partition.

    Locates the latest scan, the previous WooCommerce catalog, and the product
    links file when the caller does not supply them. Writes the change report
    plus the new / updated / image-subset manifests into a dated reports folder.

    Returns the dict produced by compare() (see its docstring), or None when a
    required input is missing (no scan, or no previous catalog to diff against).
    The returned dict includes the manifest paths so the caller can wire them
    into the downstream image + import stages.
    """
    reports_dir = str(RUNTIME_REPORTS_DIR)
    links_dirs = [str(INTERMEDIATE_DIR), reports_dir]

    if not scan_path:
        scan_path = find_latest_scan(reports_dir)
        if not scan_path:
            scan_path = find_latest_scan(str(LEGACY_APP_DIR / "reports"))
    if not scan_path:
        return None

    if not woo_path:
        woo_path = find_previous_woo_file()
    if not woo_path:
        # No previous catalog → nothing to diff. Caller treats this as "first
        # run / full".
        return None

    date_match = re.search(r"product_details_(\d{8}_\d{6})", os.path.basename(scan_path))
    date_str = date_match.group(1) if date_match else None
    if date_str:
        dated = f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"
        out_dir = str(get_dated_reports_dir(dated))
        output_path = os.path.join(out_dir, f"product_changes_{date_str}.xlsx")
    else:
        output_path = os.path.join(reports_dir, "product_changes_unknown.xlsx")

    if not links_path:
        for d in links_dirs:
            links_path = find_links_file(d)
            if links_path:
                break

    return compare(scan_path, woo_path, output_path, links_path)


# ─── اجرا ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    # default reports dir as string for backward compatibility
    reports_dir = str(RUNTIME_REPORTS_DIR)
    woo_dirs = [
        str(RUNTIME_CACHE_DIR / "import_builder" / "uploads"),
    ]
    # extracted_products.xlsx (فایل لینک محصولات) در data/intermediate تولید می‌شود
    links_dirs = [
        str(INTERMEDIATE_DIR),
        reports_dir,
    ]

    if len(sys.argv) == 1:
        scan_path = find_latest_scan(reports_dir)
        if not scan_path:
            # Legacy scans not yet migrated out of product_extraction/reports/
            scan_path = find_latest_scan(str(LEGACY_APP_DIR / "reports"))
        if not scan_path:
            print(
                "✗ Could not find product_details_YYYYMMDD_HHMMSS.xlsx in reports folder"
            )
            sys.exit(1)

        woo_path = None
        for d in woo_dirs:
            woo_path = find_latest_woo_file(d)
            if woo_path:
                break
        if not woo_path:
            print("✗ Could not find woocommerce_import_*.csv in uploads folder")
            sys.exit(1)

        date_match = re.search(r"product_details_(\d{8}_\d{6})", os.path.basename(scan_path))
        date_str = date_match.group(1) if date_match else None

        # Save in a dated subdirectory (YYYY-MM-DD) when possible to keep reports tidy
        if date_str:
            # date_str is YYYYMMDD_HHMMSS -> YYYY-MM-DD
            dated = f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"
            dated_dir = get_dated_reports_dir(dated)
            reports_dir = str(dated_dir)
            output_path = os.path.join(reports_dir, f"product_changes_{date_str}.xlsx")
        else:
            output_path = os.path.join(reports_dir, f"product_changes_unknown.xlsx")

        links_path = None
        for d in links_dirs:
            links_path = find_links_file(d)
            if links_path:
                break

        print(f"→ WooCommerce (current site): {os.path.basename(woo_path)}")
        print(f"→ New scan:                  {os.path.basename(scan_path)}")
        print(f"→ Output file:               {output_path}")
        if links_path:
            print(f"→ Product links file:        {os.path.basename(links_path)}")
        else:
            print(f"→ Product links file:        not found (links sheet will be skipped)")
        print()

        # ── هشدار ترتیب زمانی ────────────────────────────────────────
        # WooCommerce CSV باید وضعیت *قبلی* سایت باشد؛ یعنی قدیمی‌تر از اسکن جدید.
        # اگر CSV جدیدتر از اسکن باشد احتمالاً از همین اسکن ساخته شده و
        # مقایسه بی‌معنی (تقریباً بدون تغییر) خواهد بود.
        _scan_ts = re.search(r"(\d{8}_\d{6})", os.path.basename(scan_path))
        _woo_ts = re.search(r"(\d{8}_\d{6})", os.path.basename(woo_path))
        if _scan_ts and _woo_ts and _woo_ts.group(1) >= _scan_ts.group(1):
            print("⚠️  Warning: WooCommerce file is newer than (or same as) the scan.")
            print(f"    WooCommerce: {_woo_ts.group(1)}")
            print(f"    Scan:        {_scan_ts.group(1)}")
            print("    The CSV was likely generated from this scan and no meaningful changes will be found.")
            if sys.stdin.isatty():
                answer = input("    Continue? (y/N): ").strip().lower()
                if answer not in ("y", "yes"):
                    print("Cancelled.")
                    sys.exit(0)
            else:
                print("    (Non-interactive mode — continuing with the same files)")
            print()

    elif len(sys.argv) in (4, 5):
        scan_path, woo_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]
        if len(sys.argv) == 5:
            links_path = sys.argv[4]
        else:
            links_path = None
            for d in links_dirs:
                links_path = find_links_file(d)
                if links_path:
                    break

    else:
        print("Usage:")
        print(
            "  python compare_scans.py                                          <- auto mode"
        )
        print(
            "  python compare_scans.py scan.xlsx woo.csv out.xlsx                <- manual mode"
        )
        print(
            "  python compare_scans.py scan.xlsx woo.csv out.xlsx links.xlsx     <- manual mode + links"
        )
        sys.exit(1)

    compare(scan_path, woo_path, output_path, links_path)
