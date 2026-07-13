#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
برنامه پیگیری و مقایسه قیمت محصولات - نسخه پیشرفته
Product Price Tracking and Comparison Tool - Enhanced Edition

ویژگی‌های جدید:
- شناسایی محصولات حذف شده
- تولید گزارش HTML زیبا با تب‌ها
- گزارش‌های جامع‌تر
- سیستم سوابق قیمت (Price History)

تغییر:
- مسیر پوشه reports و فایل ورودی extracted_products.xlsx دیگر به
  cwd (پوشه اجرا) وابسته نیستند و نسبت به ریشه پروژه (ROOT_DIR)
  محاسبه می‌شوند. قبلاً اگر این اسکریپت از بیرون پوشه پروژه (مثلاً
  از یک .bat در پوشه بالاتر) اجرا می‌شد، یک پوشه reports جدید و
  اشتباه در محل اشتباه ساخته می‌شد بدون اینکه خطایی نشان داده شود.
"""

# Fix Windows console encoding
import sys
import codecs
if sys.platform == 'win32':
    try:
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'replace')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'replace')
    except:
        pass

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Enable direct-script execution
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# ─── Shared Excel utilities (Unit 3) ───────────────────────────────
from common.excel_utils import (
    GREEN_FILL,
    RED_FILL,
    YELLOW_FILL,
    THIN_BORDER,
    CENTER_CENTER,
    read_excel,
    excel_writer,
)
from common.date_utils import get_persian_date as _get_persian_date
from common.date_utils import gregorian_to_jalali as _gregorian_to_jalali
from common.file_registry import get_file
from common.file_utils import ensure_directory, safe_copy, find_latest_dated
from common.file_utils import find_latest_dated
from common.path_registry import INTERMEDIATE_DIR, RUNTIME_REPORTS_DIR, get_dated_reports_dir
from common.price_utils import extract_price_from_text as _extract_price_from_text
from common.price_utils import format_number as _format_number
from common.price_utils import parse_numeric_price as _parse_numeric_price
from common.text_utils import extract_product_code as _extract_product_code
from common.text_utils import extract_product_name as _extract_product_name

# ─── مسیر ریشه پروژه (مستقل از cwd) ────────────────────────────────
from common.path_registry import ROOT_DIR

# Import Price History Manager
try:
    from trackers.price_history_manager import PriceHistoryManager
    from config.history_config import HISTORY_SETTINGS
    HAS_HISTORY = True
except:
    HAS_HISTORY = False
    print("[WARNING] Price History Manager not available")
    sys.stdout.flush()


def create_reports_folder():
    """
    ایجاد فولدر reports اگر وجود ندارد (نسبت به ریشه پروژه، نه cwd)
    """
    # Default to a dated subfolder for today's date
    return get_dated_reports_dir()


def find_latest_tracking_file():
    """
    پیدا کردن آخرین فایل پیگیری (نسبت به ریشه پروژه، نه cwd)
    """
    # Prefer an explicit LATEST file in the root reports folder
    latest_file = RUNTIME_REPORTS_DIR / get_file('product_tracking_latest')
    if latest_file.exists():
        return latest_file

    # If not present, scan dated subfolders for the most recent archive
    if RUNTIME_REPORTS_DIR.exists():
        # Look for product_tracking_YYYY-MM-DD.xlsx inside immediate subfolders
        candidates = []
        for sub in RUNTIME_REPORTS_DIR.iterdir():
            if not sub.is_dir():
                continue
            candidate = find_latest_dated(sub, 'product_tracking_????-??-??.xlsx', r'product_tracking_(\d{4}-\d{2}-\d{2})\.xlsx$')
            if candidate:
                candidates.append(candidate)

        if candidates:
            # return the most recently modified among candidates
            candidates.sort(key=lambda p: p.stat().st_mtime)
            return candidates[-1]

    return None


def load_previous_data(file_path):
    """
    Load previous data
    """
    try:
        df = read_excel(file_path, sheet_name='محصولات فعلی')
        print(f"✓ Previous file loaded: {file_path.name}")
        sys.stdout.flush()
        print(f"  Previous products count: {len(df)}")
        sys.stdout.flush()
        return df
    except Exception as e:
        print(f"✗ Error loading previous file: {e}")
        sys.stdout.flush()
        return None


def process_current_data(input_file):
    """
    Process current data
    """
    print(f"\n{'='*80}")
    print("Processing new file...")
    sys.stdout.flush()
    print(f"{'='*80}")
    
    df = read_excel(input_file)
    print(f"✓ New products count: {len(df)}")
    sys.stdout.flush()
    
    # Extract information
    df['نام محصول'] = df['Product Name'].apply(extract_product_name)
    df['کد محصول'] = df['Product Name'].apply(extract_product_code)
    if 'Price' in df.columns:
        df['قیمت فعلی'] = df['Price'].apply(extract_price_from_text)
    else:
        df['قیمت فعلی'] = df['Product Name'].apply(extract_price_from_text)
    df['لینک محصول'] = df['Product URL']
    df['تاریخ بررسی'] = get_persian_date()
    df['وضعیت'] = 'نامشخص'
    
    # Reorder columns
    df = df[['No', 'نام محصول', 'کد محصول', 'قیمت فعلی', 'لینک محصول', 'تاریخ بررسی', 'وضعیت']]
    
    return df


def compare_data(current_df, previous_df):
    """
    Compare current data with previous data
    ✨ NEW: Now detects removed products!
    """
    print(f"\n{'='*80}")
    print("Comparing with previous data...")
    sys.stdout.flush()
    print(f"{'='*80}")
    
    new_products = []
    price_changes = []
    unchanged_products = []
    removed_products = []  # ✨ جدید
    
    # Helper function to parse price
    def parse_price(price):
        return _parse_numeric_price(price)
    
    # Create dictionaries
    previous_dict = {}
    current_urls = set()
    
    if previous_df is not None:
        for _, row in previous_df.iterrows():
            url = row.get('لینک محصول', row.get('Product URL', ''))
            price = row.get('قیمت فعلی', row.get('Price', None))
            code = row.get('کد محصول', '')
            previous_dict[url] = {
                'name': row.get('نام محصول', row.get('Product Name', '')),
                'price': parse_price(price),
                'code': code
            }
    
    # Compare each current product
    total = len(current_df)
    print(f"\n→ Comparing {total} products...")
    sys.stdout.flush()
    
    for idx, row in current_df.iterrows():
        url = row['لینک محصول']
        current_price = row['قیمت فعلی']
        current_urls.add(url)
        
        # Progress every 50 products
        if (idx + 1) % 50 == 0:
            print(f"  → Compared {idx + 1}/{total} products...")
            sys.stdout.flush()
        
        if url in previous_dict:
            # Product existed before
            prev_price = previous_dict[url]['price']
            
            # Check price change
            if pd.notna(current_price) and prev_price is not None:
                if current_price != prev_price:
                    # Price changed
                    change_amount = current_price - prev_price
                    change_percent = (change_amount / prev_price) * 100 if prev_price > 0 else 0
                    
                    price_changes.append({
                        'نام محصول': row['نام محصول'],
                        'کد محصول': row['کد محصول'],
                        'لینک محصول': url,
                        'قیمت قبلی': int(prev_price),
                        'قیمت جدید': int(current_price),
                        'تغییر (تومان)': int(change_amount),
                        'درصد تغییر': round(change_percent, 2),
                        'تاریخ بررسی': row['تاریخ بررسی']
                    })
                    
                    current_df.at[idx, 'وضعیت'] = 'تغییر قیمت'
                else:
                    # No change
                    unchanged_products.append(row.to_dict())
                    current_df.at[idx, 'وضعیت'] = 'بدون تغییر'
            else:
                current_df.at[idx, 'وضعیت'] = 'بدون تغییر'
        else:
            # New product
            new_products.append(row.to_dict())
            current_df.at[idx, 'وضعیت'] = 'جدید'
    
    # ✨ Find removed products (in previous but not in current)
    if previous_df is not None:
        for url, data in previous_dict.items():
            if url not in current_urls:
                removed_products.append({
                    'نام محصول': data['name'],
                    'کد محصول': data['code'],
                    'لینک محصول': url,
                    'قیمت آخرین': int(data['price']) if data['price'] else None,
                    'تاریخ حذف': get_persian_date()
                })
    
    print(f"✓ New products: {len(new_products)}")
    sys.stdout.flush()
    print(f"✓ Price changes: {len(price_changes)}")
    sys.stdout.flush()
    print(f"✓ Unchanged: {len(unchanged_products)}")
    sys.stdout.flush()
    print(f"✓ Removed products: {len(removed_products)}")  # ✨ جدید
    sys.stdout.flush()
    
    return current_df, new_products, price_changes, removed_products


def gregorian_to_jalali(g_y, g_m, g_d):
    """Compatibility wrapper for shared Jalali conversion."""
    return _gregorian_to_jalali(g_y, g_m, g_d)


def get_persian_date():
    """Compatibility wrapper for shared Persian date formatting."""
    return _get_persian_date()


def extract_price_from_text(text):
    """Compatibility wrapper for shared price extraction."""
    return _extract_price_from_text(text, min_digits=5)


def extract_product_name(text):
    """Compatibility wrapper for shared product name extraction."""
    if pd.isna(text):
        return ""
    return _extract_product_name(text)


def extract_product_code(text):
    """Compatibility wrapper for shared product code extraction."""
    if pd.isna(text):
        return ""
    return _extract_product_code(text)


def format_number(num):
    """Compatibility wrapper for shared number formatting."""
    return _format_number(num)


def style_excel_sheet(ws, sheet_type='main'):
    """
    استایل‌دهی به شیت اکسل
    """
    # رنگ‌ها
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    # new_fill / change_fill / removed_fill replaced by shared GREEN_FILL / YELLOW_FILL / RED_FILL

    # فونت
    header_font = Font(name='B Nazanin', size=12, bold=True, color="FFFFFF")
    cell_font = Font(name='B Nazanin', size=11)

    # بوردر (shared THIN_BORDER)
    thin_border = THIN_BORDER

    # استایل هدر
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = CENTER_CENTER
        cell.border = thin_border

    # استایل سلول‌ها
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = cell_font
            cell.alignment = CENTER_CENTER
            cell.border = thin_border

            # رنگ بر اساس وضعیت
            if sheet_type == 'main' and ws.cell(row=cell.row, column=7).value == 'جدید':
                cell.fill = GREEN_FILL
            elif sheet_type == 'main' and ws.cell(row=cell.row, column=7).value == 'تغییر قیمت':
                cell.fill = YELLOW_FILL
            elif sheet_type == 'removed':  # ✨ جدید
                cell.fill = RED_FILL
    
    # تنظیم عرض ستون‌ها
    column_widths = {
        'A': 8,  # No
        'B': 35, # نام محصول
        'C': 15, # کد محصول
        'D': 18, # قیمت
        'E': 45, # لینک
        'F': 15, # تاریخ
        'G': 15  # وضعیت
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ارتفاع ردیف هدر
    ws.row_dimensions[1].height = 30


def generate_html_report(current_df, new_products, price_changes, removed_products, previous_df, reports_dir):
    """
    ✨ NEW: Generate beautiful HTML report with tabs
    """
    persian_date = get_persian_date()
    date_for_filename = persian_date.replace('/', '-')
    html_file = reports_dir / f'product_tracking_report_{date_for_filename}.html'
    
    # Calculate statistics
    total_products = len(current_df)
    price_increases = sum(1 for c in price_changes if c['تغییر (تومان)'] > 0)
    price_decreases = sum(1 for c in price_changes if c['تغییر (تومان)'] < 0)
    
    # HTML content
    html_content = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>گزارش پیگیری قیمت محصولات - {persian_date}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}
        
        .header .date {{
            font-size: 1.2em;
            opacity: 0.9;
            margin-top: 10px;
        }}
        
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }}
        
        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 15px;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            transition: transform 0.3s, box-shadow 0.3s;
            border-top: 4px solid #667eea;
        }}
        
        .stat-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 8px 25px rgba(0,0,0,0.15);
        }}
        
        .stat-card .number {{
            font-size: 2.5em;
            font-weight: bold;
            color: #667eea;
            margin: 10px 0;
        }}
        
        .stat-card .label {{
            color: #666;
            font-size: 1em;
        }}
        
        .stat-card.new {{ border-top-color: #28a745; }}
        .stat-card.new .number {{ color: #28a745; }}
        
        .stat-card.changed {{ border-top-color: #ffc107; }}
        .stat-card.changed .number {{ color: #ffc107; }}
        
        .stat-card.removed {{ border-top-color: #dc3545; }}
        .stat-card.removed .number {{ color: #dc3545; }}
        
        .tabs {{
            display: flex;
            background: #e9ecef;
            padding: 0;
            overflow-x: auto;
            border-bottom: 3px solid #667eea;
        }}
        
        .tab {{
            padding: 18px 35px;
            cursor: pointer;
            background: transparent;
            border: none;
            font-size: 1.1em;
            color: #666;
            transition: all 0.3s;
            white-space: nowrap;
            border-bottom: 3px solid transparent;
            margin-bottom: -3px;
        }}
        
        .tab:hover {{
            background: rgba(102, 126, 234, 0.1);
            color: #667eea;
        }}
        
        .tab.active {{
            background: white;
            color: #667eea;
            font-weight: bold;
            border-bottom-color: #667eea;
        }}
        
        .tab-content {{
            display: none;
            padding: 30px;
            animation: fadeIn 0.5s;
        }}
        
        .tab-content.active {{
            display: block;
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; transform: translateY(10px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        
        h2 {{
            color: #333;
            border-bottom: 3px solid #667eea;
            padding-bottom: 15px;
            margin: 30px 0 20px 0;
            font-size: 1.8em;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            box-shadow: 0 2px 15px rgba(0,0,0,0.1);
            border-radius: 10px;
            overflow: hidden;
        }}
        
        th, td {{
            padding: 15px;
            text-align: center;
            border-bottom: 1px solid #e9ecef;
        }}
        
        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-weight: 600;
            font-size: 1em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        tr:hover {{
            background: #f8f9fa;
        }}
        
        tr:last-child td {{
            border-bottom: none;
        }}
        
        .price {{
            font-family: 'Courier New', monospace;
            font-weight: bold;
        }}
        
        .increase {{
            color: #dc3545;
            font-weight: bold;
        }}
        
        .decrease {{
            color: #28a745;
            font-weight: bold;
        }}
        
        .discount-badge {{
            background: #28a745;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.85em;
            margin-right: 8px;
            display: inline-block;
        }}
        
        .new-badge {{
            background: #28a745;
            color: white;
            padding: 3px 10px;
            border-radius: 15px;
            font-size: 0.8em;
        }}
        
        .removed-badge {{
            background: #dc3545;
            color: white;
            padding: 3px 10px;
            border-radius: 15px;
            font-size: 0.8em;
        }}
        
        .empty-state {{
            text-align: center;
            padding: 60px 20px;
            color: #999;
            font-size: 1.2em;
        }}
        
        .empty-state::before {{
            content: "😊";
            display: block;
            font-size: 3em;
            margin-bottom: 15px;
        }}
        
        a {{
            color: #667eea;
            text-decoration: none;
        }}
        
        a:hover {{
            text-decoration: underline;
        }}
        
        .footer {{
            background: #f8f9fa;
            padding: 20px;
            text-align: center;
            color: #666;
            border-top: 1px solid #e9ecef;
        }}
    </style>
    <script>
        function showTab(tabId) {{
            // Hide all tabs
            document.querySelectorAll('.tab-content').forEach(content => {{
                content.classList.remove('active');
            }});
            document.querySelectorAll('.tab').forEach(tab => {{
                tab.classList.remove('active');
            }});
            
            // Show selected tab
            document.getElementById(tabId).classList.add('active');
            event.currentTarget.classList.add('active');
        }}
    </script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 گزارش پیگیری قیمت محصولات</h1>
            <div class="date">تاریخ بررسی: {persian_date}</div>
        </div>
        
        <div class="stats">
            <div class="stat-card">
                <div class="label">📦 مجموع محصولات</div>
                <div class="number">{total_products:,}</div>
            </div>
            
            <div class="stat-card new">
                <div class="label">✨ محصولات جدید</div>
                <div class="number">{len(new_products):,}</div>
            </div>
            
            <div class="stat-card changed">
                <div class="label">💰 تغییرات قیمت</div>
                <div class="number">{len(price_changes):,}</div>
            </div>
            
            <div class="stat-card removed">
                <div class="label">❌ محصولات حذف شده</div>
                <div class="number">{len(removed_products):,}</div>
            </div>
            
            <div class="stat-card">
                <div class="label">⬆️ افزایش قیمت</div>
                <div class="number">{price_increases:,}</div>
            </div>
            
            <div class="stat-card">
                <div class="label">⬇️ کاهش قیمت</div>
                <div class="number">{price_decreases:,}</div>
            </div>
        </div>
        
        <div class="tabs">
            <button class="tab active" onclick="showTab('overview')">🌐 نمای کلی</button>
            <button class="tab" onclick="showTab('new')">✨ محصولات جدید</button>
            <button class="tab" onclick="showTab('changes')">💰 تغییرات قیمت</button>
            <button class="tab" onclick="showTab('removed')">❌ محصولات حذف شده</button>
            <button class="tab" onclick="showTab('all')">📋 همه محصولات</button>
        </div>
        
        <!-- Overview Tab -->
        <div id="overview" class="tab-content active">
            <h2>📊 خلاصه تغییرات</h2>
            
            <div style="background: #f8f9fa; padding: 25px; border-radius: 10px; margin: 20px 0;">
                <h3 style="color: #667eea; margin-bottom: 15px;">📈 آمار کلی</h3>
                <ul style="list-style: none; padding: 0;">
                    <li style="padding: 10px 0; border-bottom: 1px solid #ddd;">
                        <strong>مجموع محصولات:</strong> {total_products:,} عدد
                    </li>
                    <li style="padding: 10px 0; border-bottom: 1px solid #ddd;">
                        <strong>محصولات جدید:</strong> <span style="color: #28a745;">{len(new_products):,} عدد</span>
                    </li>
                    <li style="padding: 10px 0; border-bottom: 1px solid #ddd;">
                        <strong>تغییرات قیمت:</strong> <span style="color: #ffc107;">{len(price_changes):,} عدد</span>
                    </li>
                    <li style="padding: 10px 0; border-bottom: 1px solid #ddd;">
                        <strong>محصولات حذف شده:</strong> <span style="color: #dc3545;">{len(removed_products):,} عدد</span>
                    </li>
                    <li style="padding: 10px 0;">
                        <strong>تاریخ بررسی:</strong> {persian_date}
                    </li>
                </ul>
            </div>
"""
    
    # Add sample new products to overview
    if new_products:
        html_content += """
            <h3 style="color: #28a745; margin-top: 30px;">✨ نمونه محصولات جدید</h3>
            <table>
                <tr>
                    <th>نام محصول</th>
                    <th>کد</th>
                    <th>قیمت</th>
                </tr>
"""
        for product in new_products[:5]:
            price_display = format_number(product.get('قیمت فعلی', ''))
            html_content += f"""
                <tr>
                    <td>{product['نام محصول']}</td>
                    <td>{product['کد محصول']}</td>
                    <td class="price">{price_display} تومان</td>
                </tr>
"""
        html_content += "</table>"
        if len(new_products) > 5:
            html_content += f"<p style='text-align: center; color: #666; margin-top: 10px;'>و {len(new_products) - 5} محصول دیگر...</p>"
    
    # Add sample price changes to overview
    if price_changes:
        html_content += """
            <h3 style="color: #ffc107; margin-top: 30px;">💰 نمونه تغییرات قیمت</h3>
            <table>
                <tr>
                    <th>نام محصول</th>
                    <th>قیمت قبلی</th>
                    <th>قیمت جدید</th>
                    <th>تغییر</th>
                </tr>
"""
        for change in price_changes[:5]:
            direction_class = 'decrease' if change['تغییر (تومان)'] < 0 else 'increase'
            direction_symbol = '⬇️' if change['تغییر (تومان)'] < 0 else '⬆️'
            discount_badge = '<span class="discount-badge">تخفیف</span>' if change['درصد تغییر'] <= -10 else ''
            
            html_content += f"""
                <tr>
                    <td>{change['نام محصول']}</td>
                    <td class="price">{format_number(change['قیمت قبلی'])} تومان</td>
                    <td class="price">{format_number(change['قیمت جدید'])} تومان</td>
                    <td class="{direction_class}">{direction_symbol} {change['درصد تغییر']:+.1f}% {discount_badge}</td>
                </tr>
"""
        html_content += "</table>"
        if len(price_changes) > 5:
            html_content += f"<p style='text-align: center; color: #666; margin-top: 10px;'>و {len(price_changes) - 5} تغییر دیگر...</p>"
    
    html_content += """
        </div>
        
        <!-- New Products Tab -->
        <div id="new" class="tab-content">
            <h2>✨ محصولات جدید ({:,})</h2>
""".format(len(new_products))
    
    if new_products:
        html_content += """
            <table>
                <tr>
                    <th>ردیف</th>
                    <th>نام محصول</th>
                    <th>کد محصول</th>
                    <th>قیمت</th>
                    <th>لینک</th>
                </tr>
"""
        for i, product in enumerate(new_products, 1):
            price_display = format_number(product.get('قیمت فعلی', ''))
            link = product.get('لینک محصول', '#')
            html_content += f"""
                <tr>
                    <td>{i}</td>
                    <td>{product['نام محصول']}</td>
                    <td>{product['کد محصول']}</td>
                    <td class="price">{price_display} تومان</td>
                    <td><a href="{link}" target="_blank">مشاهده</a></td>
                </tr>
"""
        html_content += "</table>"
    else:
        html_content += '<div class="empty-state">محصول جدیدی یافت نشد</div>'
    
    html_content += """
        </div>
        
        <!-- Price Changes Tab -->
        <div id="changes" class="tab-content">
            <h2>💰 تغییرات قیمت ({:,})</h2>
""".format(len(price_changes))
    
    if price_changes:
        html_content += """
            <table>
                <tr>
                    <th>نام محصول</th>
                    <th>کد</th>
                    <th>قیمت قبلی</th>
                    <th>قیمت جدید</th>
                    <th>تغییر (تومان)</th>
                    <th>درصد تغییر</th>
                    <th>لینک</th>
                </tr>
"""
        for change in price_changes:
            direction_class = 'decrease' if change['تغییر (تومان)'] < 0 else 'increase'
            direction_symbol = '⬇️' if change['تغییر (تومان)'] < 0 else '⬆️'
            discount_badge = '<span class="discount-badge">تخفیف ویژه</span>' if change['درصد تغییر'] <= -10 else ''
            link = change.get('لینک محصول', '#')
            
            html_content += f"""
                <tr>
                    <td>{change['نام محصول']}</td>
                    <td>{change['کد محصول']}</td>
                    <td class="price">{format_number(change['قیمت قبلی'])} تومان</td>
                    <td class="price">{format_number(change['قیمت جدید'])} تومان</td>
                    <td class="{direction_class} price">{format_number(abs(change['تغییر (تومان)']))} تومان</td>
                    <td class="{direction_class}">{direction_symbol} {change['درصد تغییر']:+.1f}% {discount_badge}</td>
                    <td><a href="{link}" target="_blank">مشاهده</a></td>
                </tr>
"""
        html_content += "</table>"
    else:
        html_content += '<div class="empty-state">تغییر قیمتی یافت نشد</div>'
    
    # ✨ Removed Products Tab
    html_content += """
        </div>
        
        <!-- Removed Products Tab -->
        <div id="removed" class="tab-content">
            <h2>❌ محصولات حذف شده ({:,})</h2>
""".format(len(removed_products))
    
    if removed_products:
        html_content += """
            <table>
                <tr>
                    <th>ردیف</th>
                    <th>نام محصول</th>
                    <th>کد محصول</th>
                    <th>قیمت آخرین</th>
                    <th>تاریخ حذف</th>
                    <th>لینک</th>
                </tr>
"""
        for i, product in enumerate(removed_products, 1):
            price_display = format_number(product.get('قیمت آخرین', '')) if product.get('قیمت آخرین') else 'N/A'
            link = product.get('لینک محصول', '#')
            html_content += f"""
                <tr style="background: rgba(220, 53, 69, 0.05);">
                    <td>{i}</td>
                    <td>{product['نام محصول']}</td>
                    <td>{product['کد محصول']}</td>
                    <td class="price">{price_display} تومان</td>
                    <td>{product['تاریخ حذف']}</td>
                    <td><a href="{link}" target="_blank">مشاهده</a></td>
                </tr>
"""
        html_content += "</table>"
    else:
        html_content += '<div class="empty-state">محصول حذف شده‌ای یافت نشد</div>'
    
    # All Products Tab
    html_content += """
        </div>
        
        <!-- All Products Tab -->
        <div id="all" class="tab-content">
            <h2>📋 همه محصولات ({:,})</h2>
            <table>
                <tr>
                    <th>ردیف</th>
                    <th>نام محصول</th>
                    <th>کد</th>
                    <th>قیمت</th>
                    <th>وضعیت</th>
                    <th>لینک</th>
                </tr>
""".format(len(current_df))
    
    for idx, row in current_df.iterrows():
        price_display = format_number(row['قیمت فعلی'])
        status = row['وضعیت']
        status_badge = ''
        if status == 'جدید':
            status_badge = '<span class="new-badge">جدید</span>'
        elif status == 'تغییر قیمت':
            status_badge = '<span class="discount-badge">تغییر قیمت</span>'
        
        link = row['لینک محصول']
        
        html_content += f"""
                <tr>
                    <td>{row['No']}</td>
                    <td>{row['نام محصول']}</td>
                    <td>{row['کد محصول']}</td>
                    <td class="price">{price_display} تومان</td>
                    <td>{status_badge if status_badge else status}</td>
                    <td><a href="{link}" target="_blank">مشاهده</a></td>
                </tr>
"""
    
    html_content += """
            </table>
        </div>
        
        <div class="footer">
            <p>گزارش تولید شده توسط سیستم پیگیری قیمت محصولات</p>
            <p>تاریخ تولید: {}</p>
        </div>
    </div>
</body>
</html>
""".format(persian_date)
    
    # Save HTML file
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✓ HTML report saved: {html_file}")
    sys.stdout.flush()
    return html_file


def save_to_excel(current_df, new_products, price_changes, removed_products, previous_df, reports_dir):
    """
    Save data to Excel file
    ✨ Updated: Now includes removed products sheet
    """
    persian_date = get_persian_date()
    date_for_filename = persian_date.replace('/', '-')
    
    # File names in reports folder (reports_dir is a dated folder)
    latest_file = reports_dir / get_file('product_tracking_latest')
    archive_file = reports_dir / f'product_tracking_{date_for_filename}.xlsx'
    
    print(f"\n{'='*80}")
    print("Saving output files...")
    print(f"{'='*80}")
    
    # Create Excel file
    with pd.ExcelWriter(latest_file, engine='openpyxl') as writer:
        
        # شیت 1: محصولات فعلی
        df_output = current_df.copy()
        df_output['قیمت فعلی'] = df_output['قیمت فعلی'].apply(lambda x: format_number(x) if pd.notna(x) else "")
        df_output.to_excel(writer, sheet_name='محصولات فعلی', index=False)
        
        # شیت 2: محصولات جدید
        if new_products:
            df_new = pd.DataFrame(new_products)
            df_new['قیمت فعلی'] = df_new['قیمت فعلی'].apply(lambda x: format_number(x) if pd.notna(x) else "")
            df_new.to_excel(writer, sheet_name='محصولات جدید', index=False)
        else:
            pd.DataFrame({'پیام': ['محصول جدیدی یافت نشد']}).to_excel(writer, sheet_name='محصولات جدید', index=False)
        
        # شیت 3: تغییرات قیمت
        if price_changes:
            df_changes = pd.DataFrame(price_changes)
            df_changes['قیمت قبلی'] = df_changes['قیمت قبلی'].apply(format_number)
            df_changes['قیمت جدید'] = df_changes['قیمت جدید'].apply(format_number)
            df_changes['تغییر (تومان)'] = df_changes['تغییر (تومان)'].apply(lambda x: format_number(x) if x >= 0 else f"-{format_number(abs(x))}")
            df_changes['درصد تغییر'] = df_changes['درصد تغییر'].apply(lambda x: f"{x:+.2f}%")
            df_changes.to_excel(writer, sheet_name='تغییرات قیمت', index=False)
        else:
            pd.DataFrame({'پیام': ['تغییر قیمتی یافت نشد']}).to_excel(writer, sheet_name='تغییرات قیمت', index=False)
        
        # ✨ شیت 4: محصولات حذف شده (جدید)
        if removed_products:
            df_removed = pd.DataFrame(removed_products)
            df_removed['قیمت آخرین'] = df_removed['قیمت آخرین'].apply(lambda x: format_number(x) if pd.notna(x) and x else "")
            df_removed.to_excel(writer, sheet_name='محصولات حذف شده', index=False)
        else:
            pd.DataFrame({'پیام': ['محصول حذف شده‌ای یافت نشد']}).to_excel(writer, sheet_name='محصولات حذف شده', index=False)
        
        # شیت 5: تاریخچه
        if previous_df is not None:
            df_history = previous_df.copy()
            if 'قیمت فعلی' in df_history.columns:
                df_history['قیمت فعلی'] = df_history['قیمت فعلی'].apply(lambda x: format_number(x) if pd.notna(x) else "")
            df_history.to_excel(writer, sheet_name='تاریخچه', index=False)
        else:
            pd.DataFrame({'پیام': ['این اولین اجراست - تاریخچه‌ای وجود ندارد']}).to_excel(writer, sheet_name='تاریخچه', index=False)
    
    # استایل‌دهی
    wb = openpyxl.load_workbook(latest_file)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            sheet_type = 'removed' if sheet_name == 'محصولات حذف شده' else ('main' if sheet_name == 'محصولات فعلی' else 'other')
            style_excel_sheet(ws, sheet_type=sheet_type)
    
    wb.save(latest_file)
    print(f"✓ LATEST file saved: {latest_file}")
    sys.stdout.flush()
    
    # Copy to archive
    import shutil
    shutil.copy2(latest_file, archive_file)
    print(f"✓ Archive file saved: {archive_file}")
    sys.stdout.flush()
    
    # Also keep/update a root-level LATEST copy for backward compatibility
    root_latest = RUNTIME_REPORTS_DIR / get_file('product_tracking_latest')
    try:
        safe_copy(latest_file, root_latest, overwrite=True)
        print(f"✓ Root LATEST updated: {root_latest}")
        sys.stdout.flush()
    except Exception:
        pass

    return latest_file, archive_file


def main():
    """
    Main function
    """
    print("""
    ╔════════════════════════════════════════════════════════╗
    ║   Product Price Tracking & Comparison Tool v2.0        ║
    ║                Enhanced Edition                        ║
    ║   ✨ Detects removed products and generates HTML reports ║
    ╚════════════════════════════════════════════════════════╝
    """)
    
    # ✨ ایجاد فولدر reports (نسبت به ریشه پروژه، نه cwd)
    reports_dir = create_reports_folder()
    print(f"📁 Reports folder: {reports_dir.absolute()}\n")
    
    # ورودی نسبت به ریشه پروژه (نه cwd)
    input_file = INTERMEDIATE_DIR / get_file('extracted_products')
    
    # Check input file exists
    if not Path(input_file).exists():
        print(f"✗ Error: File {input_file} not found!")
        sys.stdout.flush()
        return
    
    print(f"✓ Input file found: {input_file}")
    sys.stdout.flush()
    
    # Find previous file in reports folder
    previous_file = find_latest_tracking_file()
    
    if previous_file:
        print(f"✓ Previous file found: {previous_file}")
        sys.stdout.flush()
        previous_df = load_previous_data(previous_file)
    else:
        print("ℹ First run - No previous file found")
        sys.stdout.flush()
        previous_df = None
    
    # Process current data
    current_df = process_current_data(input_file)
    
    # Initialize Price History Manager
    history_manager = None
    if HAS_HISTORY and HISTORY_SETTINGS['enable_history']:
        history_manager = PriceHistoryManager(reports_dir)
        print("\n[OK] Price History Manager initialized")
        sys.stdout.flush()
        
        # Show history statistics
        stats = history_manager.get_statistics()
        if stats['total_records'] > 0:
            print(f"    -> Total history records: {stats['total_records']}")
            print(f"    -> Products tracked: {stats['total_products']}")
            print(f"    -> Date range: {stats['date_range']}")
            sys.stdout.flush()
    
    # Compare
    if previous_df is not None:
        current_df, new_products, price_changes, removed_products = compare_data(current_df, previous_df)
    else:
        new_products = current_df.to_dict('records')
        price_changes = []
        removed_products = []
        current_df['وضعیت'] = 'جدید'
    
    # Add to Price History
    if history_manager:
        persian_date = get_persian_date()
        history_records = []
        
        # Add price changes
        for change in price_changes:
            record = history_manager.add_price_change(
                sku=str(change.get('کد محصول', '') or change.get('sku', '')),
                product_name=change.get('نام محصول', ''),
                old_price=change.get('قیمت قبلی', 0),
                new_price=change.get('قیمت جدید', 0),
                change_type='افزایش' if change.get('تغییر (تومان)', 0) > 0 else 'کاهش',
                persian_date=persian_date
            )
            history_records.append(record)
        
        # Add new products
        for new_prod in new_products:
            record = history_manager.add_price_change(
                sku=str(new_prod.get('کد محصول', '') or new_prod.get('sku', '')),
                product_name=new_prod.get('نام محصول', ''),
                old_price=0,
                new_price=new_prod.get('قیمت فعلی', 0),
                change_type='جدید',
                persian_date=persian_date
            )
            history_records.append(record)
        
        # Save history
        if history_records:
            history_manager.save_history(history_records)
        
        # Create daily snapshot
        history_manager.create_daily_snapshot(current_df)
    
    # Save Excel to reports folder
    latest_file, archive_file = save_to_excel(current_df, new_products, price_changes, removed_products, previous_df, reports_dir)
    
    # ✨ Generate HTML Report in reports folder
    html_file = generate_html_report(current_df, new_products, price_changes, removed_products, previous_df, reports_dir)
    
    # Summary
    print(f"\n{'='*80}")
    print("SUMMARY:")
    print(f"{'='*80}")
    print(f"📊 Total products: {len(current_df)}")
    sys.stdout.flush()
    print(f"🆕 New products: {len(new_products)}")
    sys.stdout.flush()
    print(f"💰 Price changes: {len(price_changes)}")
    print(f"❌ Removed products: {len(removed_products)}")  # ✨ جدید
    sys.stdout.flush()
    
    # Price History Summary
    if history_manager:
        stats = history_manager.get_statistics()
        print(f"\n📈 Price History:")
        print(f"  • Total records: {stats['total_records']}")
        print(f"  • Products tracked: {stats['total_products']}")
        print(f"  • Price increases: {stats.get('price_increases', 0)}")
        print(f"  • Price decreases: {stats.get('price_decreases', 0)}")
        sys.stdout.flush()
    
    print(f"\n📅 Check date: {get_persian_date()}")
    print(f"\n✓ Output files in '{reports_dir}' folder:")
    sys.stdout.flush()
    print(f"  • {latest_file.name}")
    print(f"  • {archive_file.name}")
    print(f"  • {html_file.name}")  # ✨ جدید
    if history_manager:
        print(f"  • {HISTORY_SETTINGS['history_filename']}")
    print(f"\n{'='*80}")
    print("✓ Processing completed successfully!")
    sys.stdout.flush()
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()
